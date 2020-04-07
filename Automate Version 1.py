import openpyxl as xl
from openpyxl.styles import Font, colors

'''
Code for By Alphabetical sheet
'''


# makes an array taken from existing anime list
def make_array(sheet):
    a = []
    for column in range(1, 7):
        for row in range(1, 22):
            cell = sheet.cell(row, column)
            if cell.value == None:
                break
            a.append(cell.value)

    if sheet == wb['By Favorite']:
        for i in range(len(a)):
            j = 0
            m = 0
            while a[i][j] != " ":
                j += 1
                m += 1
            m += 2
            a[i] = a[i][m:]

    return a


# turns array into a matrix
def order_array(sheet, a):
    a2 = []
    end = 21
    start = 0
    for i in range(((len(a) // 21) + 1)):
        a2.append(a[start:end])
        if end + 21 <= len(a):
            end = end + 21
        else:
            end = len(a)
        start = start + 21
    return a2


# prints the matrix in a sheet
def print_array(sheet, a):
    if sheet == wb['By Favorite']:
        i = 1
        c = 1
        for column in range(len(a)):
            r = 1
            for row in range(len(a[column])):
                cell = sheet.cell(r, c)
                cell.value = str(i) + '.  ' + a[column][row]
                i += 1
                r += 1
            c += 1
    else:
        c = 1
        for column in range(len(a)):
            r = 1
            for row in range(len(a[column])):
                cell = sheet.cell(r, c)
                cell.value = a[column][row]
                r += 1
            c += 1


# adds element to array
def add(a):
    flag = True
    while flag:
        p = input("Which anime would you like to add? ").strip()
        if p == "close" or p == "" or p == "no":
            flag = False
        else:
            a.append(p)
            print("Anime has been added")


# deletes element from array
def delete(a):
    flag = True
    while flag:
        p = input("Which anime would you like to remove? ").strip()
        if p == "close" or p == "" or p == "no":
            flag = False
        else:
            try:
                a.remove(p)
                print("Anime has been removed")
            except ValueError:
                print("Anime is not in list")


# sets given cells to None type
def clean_slate(sheet):
    FT = Font(name='Calibri',
              color=colors.BLACK,
              bold=False,
              )

    for column in range(1, 7):
        for row in range(1, 22):
            cell = sheet.cell(row, column)
            cell.font = FT
            cell.value = None


def whole_list(c):
    p = 'asdfs'
    j = -1
    try:
        o = int(input('Where would you like to start: '))
        for i in range(o, len(c)):
            if p == 'no' or p == 'close' or p == '':
                break
            if j == len(c) -1:
                break
            j = i
            flag = True
            while flag:
                if j != 0:
                    print('Above:', str(j) + '.', c[j - 1])
                if j != len(c) - 1:
                    print('Below:', str(j + 2) + '.', c[j + 1])

                try:
                    p = input('Where would you like to place ' + str(j + 1) + ". " + str(
                        c[j]) + ' [u]p,[d]own,[s]tay: ').lower().strip()
                except ValueError:
                    print('Rank is out of index')

                if p == 'u':
                    c[j], c[j - 1] = c[j - 1], c[j]
                    j -= 1
                    if j == len(c) - 2:
                        flag = False
                        p = 'close'
                elif p == 'd':
                    c[j], c[j + 1] = c[j + 1], c[j]
                    j += 1
                    if j == len(c):
                        flag = False
                        p = 'close'
                elif p == 's' or p == 'no' or p == 'close' or p == '':
                    flag = False
                else:
                    print('Enter [u]p,[d]own,[s]tay')
    except ValueError or UnboundLocalError:
        print('You have to enter a number')

    return c


def change_rank(c):
    flag = True
    while flag:
        p = input('Which anime would you like to switch: ')

        if p == "close" or p == "" or p == "no":
            flag = False
        elif p not in c:
            print('Anime is not in list')
        else:
            if p in c:
                try:
                    d = int(input('What rank would you like to give it: ')) - 1
                    c.remove(p)
                    c.insert(d,p)
                    print('Rank has been updated')
                except ValueError:
                    print('Given value is not a rank')
                try:
                    pass
                except ValueError:
                    print('RANK OUT OF RANGE')

    return c


def rank_unordered_list(a):
    d = [a[0]]
    flag = True
    for i in range(len(a) - 1):
        if not flag:
            break
        for j in range(len(d)):
            print(str(j + 1) + '.  ' + d[j])
        print("")

        try:
            u = int(input('Where would you like to place ' + str(a[i + 1]) + ': ')) - 1
            d.insert(u, a[j + 1])
        except ValueError:
            flag = False

    return d


wb = xl.load_workbook('List of Animes.xlsx')
sheet = wb['By Alphabetical']
a = make_array(sheet)

# user input
f = True
while f:
    m = input("Enter 1 to add or 2 to delete: ")
    if m == "close" or m == "" or m == "no":
        f = False
    elif m == "1":
        add(a)
    elif m == "2":
        delete(a)

# calls functions
clean_slate(sheet)
a.sort()
b = order_array(sheet, a)
print_array(sheet, b)

# sets cell number for Number of Animes Watched
col = len(b)
if len(b[-1]) >= 20:
    col = col + 1
    row = 2
else:
    row = len(b[-1]) + 2

sheet = wb['By Alphabetical']
cell = sheet.cell(row, col)

ft = Font(name='Dubai',
          color=colors.RED,
          bold=True,
          )

cell.font = ft
cell.value = 'Animes Watched:  ' + str(len(a))

'''
Code for By Favorite
'''
sheet = wb['By Favorite']
c = make_array(sheet)
m = True
d = ''
while m:
    k = input('Press 1 to compare every anime or 2 to update rank of anime or 3 to rank from scratch: ')
    if k == "close" or k == "" or k == "no":
        m = False
    elif k == "1":
        d = whole_list(c)
    elif k == "2":
        d = change_rank(c)
    elif k == "3":
        d = rank_unordered_list(a)

clean_slate(sheet)
b = order_array(sheet, c)

print_array(sheet, b)

col = len(b)
if len(b[-1]) >= 20:
    col = col + 1
    row = 2
else:
    row = len(b[-1]) + 2

sheet = wb['By Favorite']
cell = sheet.cell(row, col)

ft = Font(name='Dubai',
          color=colors.RED,
          bold=True,
          )

cell.font = ft
cell.value = 'Animes Watched:  ' + str(len(c))

wb.save('List of Animes.xlsx')
